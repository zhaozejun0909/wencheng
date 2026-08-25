from __future__ import annotations

import csv
import json
import sqlite3
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from wen.models import (
    CollectionResult,
    JobStatus,
    ProductSelectionMode,
    QueryGroup,
    QueryPlan,
    utc_now,
)

_CHINA_TZ = ZoneInfo("Asia/Shanghai")


def _is_catalog_result(result: CollectionResult) -> bool:
    return result.selection_mode == ProductSelectionMode.PRECISE_CATALOG


def _result_scope_name(result: CollectionResult) -> str:
    """Return the user-facing query scope, which may be a store or all favorites."""
    if result.selection_mode == ProductSelectionMode.FAVORITES:
        return "我的收藏"
    if result.selection_mode == ProductSelectionMode.PRECISE:
        return "商品 ID 查询"
    return (
        result.store.name
        if result.store and result.store.name
        else (result.requested_store_name or result.requested_sec_shop_id or result.keyword)
    )


def format_china_time(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(_CHINA_TZ).strftime("%Y-%m-%d %H:%M:%S")


class DataStore:
    def __init__(self, database_path: Path) -> None:
        database_path.parent.mkdir(parents=True, exist_ok=True)
        self.path = database_path
        self._init_schema()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        return connection

    def _init_schema(self) -> None:
        with self._connect() as connection:
            connection.executescript(
                """
                PRAGMA journal_mode=WAL;
                CREATE TABLE IF NOT EXISTS jobs (
                    id TEXT PRIMARY KEY,
                    keyword TEXT NOT NULL,
                    backend TEXT NOT NULL,
                    status TEXT NOT NULL,
                    state TEXT NOT NULL,
                    started_at TEXT,
                    finished_at TEXT,
                    result_json TEXT,
                    error TEXT,
                    history_excluded INTEGER NOT NULL DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS store_snapshots (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_id TEXT NOT NULL,
                    keyword TEXT NOT NULL,
                    name TEXT,
                    douyin_id TEXT,
                    category TEXT,
                    followers INTEGER,
                    product_count INTEGER,
                    description TEXT,
                    collected_at TEXT NOT NULL,
                    raw_json TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS product_snapshots (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    job_id TEXT NOT NULL,
                    store_snapshot_id INTEGER,
                    title TEXT NOT NULL,
                    price REAL,
                    original_price REAL,
                    displayed_sales INTEGER,
                    displayed_sales_raw TEXT,
                    rating REAL,
                    review_count INTEGER,
                    product_id TEXT,
                    source_url TEXT,
                    position INTEGER,
                    collected_at TEXT NOT NULL,
                    raw_json TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS query_groups (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    store_name TEXT NOT NULL,
                    config_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS query_plans (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    config_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                """
            )
            columns = {
                row["name"]
                for row in connection.execute("PRAGMA table_info(jobs)").fetchall()
            }
            if "query_run_id" not in columns:
                connection.execute("ALTER TABLE jobs ADD COLUMN query_run_id TEXT")
            if "history_excluded" not in columns:
                connection.execute(
                    "ALTER TABLE jobs ADD COLUMN history_excluded INTEGER NOT NULL DEFAULT 0"
                )
            # Older versions recorded precise-catalog discovery as a regular job.
            # Keep those rows available for polling/debugging, but exclude them
            # from query history and mark them for the one-time cleanup below.
            rows = connection.execute(
                "SELECT id, result_json FROM jobs WHERE history_excluded=0 AND result_json IS NOT NULL"
            ).fetchall()
            for row in rows:
                try:
                    payload = json.loads(row["result_json"])
                except (TypeError, ValueError):
                    continue
                if payload.get("selection_mode") == ProductSelectionMode.PRECISE_CATALOG.value:
                    connection.execute(
                        "UPDATE jobs SET history_excluded=1 WHERE id=?", (row["id"],)
                    )

    def create_job(self, result: CollectionResult) -> None:
        with self._connect() as connection:
            connection.execute(
                "INSERT OR REPLACE INTO jobs(id, keyword, backend, status, state, started_at, query_run_id, history_excluded) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    result.job_id,
                    result.keyword,
                    result.backend,
                    result.status.value,
                    result.state,
                    result.started_at.isoformat(),
                    result.query_run_id,
                    int(_is_catalog_result(result)),
                ),
            )

    def update_job_state(self, job_id: str, status: JobStatus, state: str) -> None:
        """Persist live progress so polling clients can show the current stage."""
        with self._connect() as connection:
            connection.execute(
                "UPDATE jobs SET status=?, state=? WHERE id=?",
                (status.value, state, job_id),
            )

    def save_result(self, result: CollectionResult) -> None:
        payload = result.model_dump_json()
        with self._connect() as connection:
            connection.execute(
                "UPDATE jobs SET status=?, state=?, finished_at=?, result_json=?, error=?, query_run_id=?, history_excluded=? WHERE id=?",
                (
                    result.status.value,
                    result.state,
                    result.finished_at.isoformat() if result.finished_at else None,
                    payload,
                    "\n".join(result.errors) if result.errors else None,
                    result.query_run_id,
                    int(_is_catalog_result(result)),
                    result.job_id,
                ),
            )
            # Catalog discovery is a temporary picker data source, not a
            # historical collection. Its result_json remains available for
            # polling, but it must not create store/product snapshots.
            if not _is_catalog_result(result) and (
                result.store or result.selection_mode == ProductSelectionMode.FAVORITES
            ):
                store = result.store
                scope_raw = (
                    store.model_dump_json()
                    if store
                    else json.dumps(
                        {"scope": "favorites", "name": "我的收藏"}, ensure_ascii=False
                    )
                )
                store_cursor = connection.execute(
                    """INSERT INTO store_snapshots
                    (job_id, keyword, name, douyin_id, category, followers, product_count, description, collected_at, raw_json)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        result.job_id,
                        result.keyword,
                        store.name if store else None,
                        store.douyin_id if store else None,
                        store.category if store else None,
                        store.followers if store else None,
                        store.product_count if store else len(result.products),
                        store.description if store else "我的收藏商品",
                        (result.finished_at or utc_now()).isoformat(),
                        scope_raw,
                    ),
                )
                store_id = store_cursor.lastrowid
                for product in result.products:
                    connection.execute(
                        """INSERT INTO product_snapshots
                        (job_id, store_snapshot_id, title, price, original_price, displayed_sales, displayed_sales_raw,
                         rating, review_count, product_id, source_url, position, collected_at, raw_json)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            result.job_id,
                            store_id,
                            product.title,
                            product.price,
                            product.original_price,
                            product.displayed_sales,
                            product.displayed_sales_raw,
                            product.rating,
                            product.review_count,
                            product.product_id,
                            product.source_url,
                            product.position,
                            (result.finished_at or utc_now()).isoformat(),
                            product.model_dump_json(),
                        ),
                    )

    def get_result(self, job_id: str) -> CollectionResult | None:
        with self._connect() as connection:
            row = connection.execute("SELECT result_json FROM jobs WHERE id=?", (job_id,)).fetchone()
        if not row or not row["result_json"]:
            return None
        return CollectionResult.model_validate_json(row["result_json"])

    def get_results(self, job_ids: list[str]) -> list[CollectionResult]:
        """Load completed results in the caller's order for a combined query run."""
        results: list[CollectionResult] = []
        for job_id in job_ids:
            result = self.get_result(job_id)
            if result is None:
                raise ValueError(f"找不到任务：{job_id}")
            results.append(result)
        return results

    def get_job_status(self, job_id: str) -> CollectionResult | None:
        """返回尚未完成任务的最小状态，供 Web 轮询使用。"""
        with self._connect() as connection:
            row = connection.execute(
                "SELECT id, keyword, backend, status, state, started_at, finished_at, error "
                "FROM jobs WHERE id=?",
                (job_id,),
            ).fetchone()
        if not row:
            return None
        status = JobStatus(row["status"])
        return CollectionResult(
            job_id=row["id"],
            status=status,
            backend=row["backend"],
            keyword=row["keyword"],
            started_at=row["started_at"] or utc_now(),
            finished_at=row["finished_at"],
            errors=[row["error"]] if row["error"] else [],
            state=row["state"],
        )

    def save_query_group(self, group: QueryGroup) -> QueryGroup:
        payload = group.model_dump_json()
        with self._connect() as connection:
            connection.execute(
                """INSERT OR REPLACE INTO query_groups
                (id, name, store_name, config_json, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)""",
                (
                    group.id,
                    group.name,
                    group.store_name or "",
                    payload,
                    group.created_at.isoformat(),
                    group.updated_at.isoformat(),
                ),
            )
        return group

    def get_query_group(self, group_id: str) -> QueryGroup | None:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT config_json FROM query_groups WHERE id=?", (group_id,)
            ).fetchone()
        return QueryGroup.model_validate_json(row["config_json"]) if row else None

    def list_query_groups(self) -> list[QueryGroup]:
        with self._connect() as connection:
            rows = connection.execute(
                # 查询条件按创建顺序展示，最新保存的自然排在最下面；编辑或重命名不改变位置。
                "SELECT config_json FROM query_groups ORDER BY created_at ASC, id ASC"
            ).fetchall()
        return [QueryGroup.model_validate_json(row["config_json"]) for row in rows]

    def delete_query_group(self, group_id: str) -> bool:
        with self._connect() as connection:
            cursor = connection.execute("DELETE FROM query_groups WHERE id=?", (group_id,))
        return cursor.rowcount > 0

    def save_query_plan(self, plan: QueryPlan) -> QueryPlan:
        with self._connect() as connection:
            connection.execute(
                """INSERT OR REPLACE INTO query_plans
                (id, name, config_json, created_at, updated_at) VALUES (?, ?, ?, ?, ?)""",
                (
                    plan.id,
                    plan.name,
                    plan.model_dump_json(),
                    plan.created_at.isoformat(),
                    plan.updated_at.isoformat(),
                ),
            )
        return plan

    def get_query_plan(self, plan_id: str) -> QueryPlan | None:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT config_json FROM query_plans WHERE id=?", (plan_id,)
            ).fetchone()
        return QueryPlan.model_validate_json(row["config_json"]) if row else None

    def list_query_plans(self) -> list[QueryPlan]:
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT config_json FROM query_plans ORDER BY updated_at DESC"
            ).fetchall()
        return [QueryPlan.model_validate_json(row["config_json"]) for row in rows]

    def count_jobs(self) -> int:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT COUNT(*) AS count FROM jobs WHERE history_excluded=0"
            ).fetchone()
        return int(row["count"] if row else 0)

    @staticmethod
    def _job_item_from_row(row: sqlite3.Row) -> dict[str, object]:
        item: dict[str, object] = {
            key: row[key]
            for key in (
                "id", "keyword", "backend", "status", "state", "started_at", "finished_at",
                "query_run_id",
            )
        }
        if row["result_json"]:
            try:
                result = CollectionResult.model_validate_json(row["result_json"])
            except ValueError:
                result = None
            if result:
                item.update(
                    {
                        "store_name": _result_scope_name(result),
                        "query_group_name": result.query_group_name,
                        "product_count": len(result.products),
                        "warning_count": len(result.warnings),
                        "error_count": len(result.errors),
                    }
                )
        return item

    def list_jobs(self, limit: int = 5, offset: int = 0) -> list[dict[str, object]]:
        limit = max(1, int(limit))
        offset = max(0, int(offset))
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT id, keyword, backend, status, state, started_at, finished_at, query_run_id, result_json "
                "FROM jobs WHERE history_excluded=0 ORDER BY started_at DESC LIMIT ? OFFSET ?",
                (limit, offset),
            ).fetchall()
        return [self._job_item_from_row(row) for row in rows]

    def count_query_runs(self) -> int:
        with self._connect() as connection:
            row = connection.execute(
                "SELECT COUNT(DISTINCT COALESCE(NULLIF(query_run_id, ''), id)) AS count "
                "FROM jobs WHERE history_excluded=0"
            ).fetchone()
        return int(row["count"] if row else 0)

    def query_run_jobs(self, run_id: str) -> list[dict[str, str]]:
        """Return the persisted jobs belonging to one user-visible query run."""
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT id, status
                FROM jobs
                WHERE history_excluded=0
                  AND (
                    query_run_id=?
                    OR ((query_run_id IS NULL OR query_run_id='') AND id=?)
                  )
                ORDER BY started_at ASC, id ASC
                """,
                (run_id, run_id),
            ).fetchall()
        return [{"id": str(row["id"]), "status": str(row["status"])} for row in rows]

    def list_query_runs(self, limit: int = 5, offset: int = 0) -> list[dict[str, object]]:
        limit = max(1, int(limit))
        offset = max(0, int(offset))
        with self._connect() as connection:
            run_rows = connection.execute(
                """
                SELECT COALESCE(NULLIF(query_run_id, ''), id) AS run_id,
                       MAX(started_at) AS latest_started_at
                FROM jobs
                WHERE history_excluded=0
                GROUP BY COALESCE(NULLIF(query_run_id, ''), id)
                ORDER BY latest_started_at DESC
                LIMIT ? OFFSET ?
                """,
                (limit, offset),
            ).fetchall()
            runs: list[dict[str, object]] = []
            for run_row in run_rows:
                run_id = run_row["run_id"]
                job_rows = connection.execute(
                    """
                    SELECT id, keyword, backend, status, state, started_at, finished_at,
                           query_run_id, result_json
                    FROM jobs
                    WHERE history_excluded=0
                      AND (query_run_id=? OR (query_run_id IS NULL AND id=?))
                    ORDER BY started_at ASC, id ASC
                    """,
                    (run_id, run_id),
                ).fetchall()
                jobs = [self._job_item_from_row(row) for row in job_rows]
                if not jobs:
                    continue
                statuses = [str(job["status"]) for job in jobs]
                if "running" in statuses:
                    status = "running"
                elif "queued" in statuses:
                    status = "queued"
                elif "failed" in statuses:
                    status = "failed"
                elif "paused" in statuses:
                    status = "paused"
                else:
                    status = "succeeded"
                started_values = [str(job["started_at"]) for job in jobs if job["started_at"]]
                finished_values = [str(job["finished_at"]) for job in jobs if job["finished_at"]]
                groups = [
                    {
                        "job_id": job["id"],
                        "store_name": job.get("store_name") or job["keyword"],
                        "query_group_name": job.get("query_group_name") or "未命名查询条件",
                        "status": job["status"],
                        "product_count": job.get("product_count"),
                        "warning_count": job.get("warning_count", 0),
                        "error_count": job.get("error_count", 0),
                    }
                    for job in jobs
                ]
                runs.append(
                    {
                        "id": run_id,
                        "status": status,
                        "backend": jobs[0]["backend"],
                        "started_at": min(started_values) if started_values else None,
                        "finished_at": max(finished_values) if len(finished_values) == len(jobs) else None,
                        "job_ids": [str(job["id"]) for job in jobs],
                        "groups": groups,
                    }
                )
        return runs

    def delete_catalog_jobs(self) -> int:
        """Delete legacy catalog-only jobs and their persisted snapshots."""
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT id FROM jobs WHERE history_excluded=1"
            ).fetchall()
            job_ids = [row["id"] for row in rows]
            if not job_ids:
                return 0
            placeholders = ",".join("?" for _ in job_ids)
            connection.execute(
                f"DELETE FROM product_snapshots WHERE job_id IN ({placeholders})", job_ids
            )
            connection.execute(
                f"DELETE FROM store_snapshots WHERE job_id IN ({placeholders})", job_ids
            )
            connection.execute(f"DELETE FROM jobs WHERE id IN ({placeholders})", job_ids)
            return len(job_ids)

    def delete_job(self, job_id: str) -> bool:
        """Remove one job and all snapshots created for it.

        A user-stopped task is intentionally not retained as history.  The caller
        removes its evidence directory separately because that path is outside
        SQLite and may still be in use until the device worker has unwound.
        """
        with self._connect() as connection:
            connection.execute("DELETE FROM product_snapshots WHERE job_id=?", (job_id,))
            connection.execute("DELETE FROM store_snapshots WHERE job_id=?", (job_id,))
            cursor = connection.execute("DELETE FROM jobs WHERE id=?", (job_id,))
            return cursor.rowcount > 0

    def export_csv(self, job_id: str, destination: Path) -> Path:
        result = self.get_result(job_id)
        if result is None:
            raise ValueError(f"找不到任务：{job_id}")
        destination.parent.mkdir(parents=True, exist_ok=True)
        with destination.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(
                file,
                fieldnames=[
                    "job_id", "store_name", "keyword", "position", "title", "price",
                    "displayed_sales", "displayed_sales_raw", "collected_at", "evidence_dir",
                ],
            )
            writer.writeheader()
            for product in result.products:
                writer.writerow(
                    {
                        "job_id": result.job_id,
                        "store_name": _result_scope_name(result),
                        "keyword": result.keyword,
                        "position": product.position,
                        "title": product.title,
                        "price": product.price,
                        "displayed_sales": product.displayed_sales,
                        "displayed_sales_raw": product.displayed_sales_raw,
                        "collected_at": format_china_time(result.finished_at),
                        "evidence_dir": result.evidence_dir,
                    }
                )
        return destination

    def export_xlsx(self, job_id: str, destination: Path) -> Path:
        return self.export_xlsx_batch([job_id], destination)

    def export_xlsx_batch(self, job_ids: list[str], destination: Path) -> Path:
        """Export one query run, with one clearly separated section per condition."""
        if not job_ids:
            raise ValueError("至少需要一个查询任务")
        results = self.get_results(job_ids)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError as exc:
            raise RuntimeError("XLSX 导出需要 openpyxl，请运行 uv sync") from exc
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "查询结果"
        sheet.freeze_panes = "A6"
        header_fill = PatternFill("solid", fgColor="1F2937")
        section_fill = PatternFill("solid", fgColor="DBEAFE")
        warning_fill = PatternFill("solid", fgColor="FEF2F2")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        thin_gray = Side(style="thin", color="D1D5DB")
        table_border = Border(bottom=thin_gray)
        detail_headers = ["商品名称", "当前价", "销量", "原始销量", "采集时间"]
        run_started = min(
            (result.started_at for result in results),
            default=None,
        )
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        run_time_cell = sheet.cell(1, 1, f"查询时间：{format_china_time(run_started) or '—'}")
        run_time_cell.font = bold_font
        run_time_cell.fill = section_fill
        current_row = 3
        first_header_row: int | None = None
        for index, result in enumerate(results):
            if index:
                current_row += 2
            store_name = _result_scope_name(result)
            query_name = result.query_group_name or "未命名查询条件"
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            title_cell = sheet.cell(current_row, 1, store_name)
            title_cell.font = Font(bold=True, size=14, color="1F2937")
            title_cell.fill = section_fill
            current_row += 1
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            condition_cell = sheet.cell(current_row, 1, f"查询条件：{query_name}")
            condition_cell.font = bold_font
            condition_cell.fill = section_fill
            current_row += 1
            missing = "、".join(result.missing_product_titles) or "无"
            metadata = (
                f"状态：{result.status.value}｜商品数：{len(result.products)}｜"
                f"精准未命中：{len(result.missing_product_titles)}｜"
                f"收藏失效排除：{len(result.invalid_favorite_titles)}"
            )
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            metadata_cell = sheet.cell(current_row, 1, metadata)
            metadata_cell.font = bold_font
            if result.missing_product_titles:
                metadata_cell.fill = warning_fill
            current_row += 1
            diagnostics = []
            diagnostics.extend(f"错误：{value}" for value in result.errors)
            diagnostics.extend(f"提示：{value}" for value in result.warnings)
            if result.missing_product_titles:
                diagnostics.append(f"精准未命中商品（可能已改名或下架）：{missing}")
            if result.invalid_favorite_titles:
                diagnostics.append(
                    "收藏失效商品（已自动排除）：" + "、".join(result.invalid_favorite_titles)
                )
            if diagnostics:
                sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                diagnostic_cell = sheet.cell(current_row, 1, "\n".join(diagnostics))
                diagnostic_cell.alignment = Alignment(wrap_text=True, vertical="top")
                diagnostic_cell.fill = warning_fill
                current_row += 1
            if first_header_row is None:
                first_header_row = current_row
            for column, value in enumerate(detail_headers, start=1):
                cell = sheet.cell(current_row, column, value)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            current_row += 1
            collected_at = format_china_time(result.finished_at)
            for product in result.products:
                values = [
                    product.title,
                    product.price,
                    product.displayed_sales,
                    product.displayed_sales_raw,
                    collected_at,
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(current_row, column, value)
                    cell.border = table_border
                    cell.alignment = Alignment(wrap_text=column in {1, 4, 5}, vertical="top")
                    if column == 2 and value is not None:
                        cell.font = Font(color="DC2626", bold=True)
                current_row += 1

        widths = {1: 44, 2: 12, 3: 14, 4: 18, 5: 26}
        for column, width in widths.items():
            sheet.column_dimensions[chr(64 + column)].width = width
        if first_header_row is not None:
            sheet.auto_filter.ref = f"A{first_header_row}:E{max(first_header_row, current_row - 1)}"

        summary = workbook.create_sheet("汇总")
        summary_headers = [
            "查询范围", "查询条件", "状态", "商品数量", "精准未命中商品", "收藏失效商品", "错误", "提示", "证据目录"
        ]
        summary.append(summary_headers)
        for cell in summary[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for result in results:
            store_name = _result_scope_name(result)
            summary.append(
                [
                    store_name,
                    result.query_group_name or "未命名查询条件",
                    result.status.value,
                    len(result.products),
                    "、".join(result.missing_product_titles) or "无",
                    "、".join(result.invalid_favorite_titles) or "无",
                    "；".join(result.errors),
                    "；".join(result.warnings),
                    result.evidence_dir,
                ]
            )
        for row in summary.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = table_border
            if row[4].value != "无" or row[5].value != "无":
                for cell in row:
                    cell.fill = warning_fill
        for column, width in enumerate((24, 28, 14, 12, 42, 42, 42, 60, 56), start=1):
            summary.column_dimensions[chr(64 + column)].width = width
        summary.freeze_panes = "A2"
        summary.auto_filter.ref = f"A1:I{max(1, summary.max_row)}"
        workbook.save(destination)
        return destination
