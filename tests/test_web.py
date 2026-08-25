import sqlite3
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from fixture_device import FixtureDevice
from openpyxl import load_workbook
from PIL import Image

from wen.config import Settings
from wen.storage import DataStore
from wen.web.app import _extract_product_id_from_url, create_app


def test_product_id_can_be_extracted_from_share_link(tmp_path: Path) -> None:
    product_id = "3583028748435489967"
    resolved_url = (
        "https://haohuo.jinritemai.com/ecommerce/trade/detail/index.html?"
        "detail_schema=sslocal%3A%2F%2Fec_goods_detail%3Fproduct_id%3D"
        f"{product_id}%26promotion_id%3D{product_id}&id={product_id}"
    )
    assert _extract_product_id_from_url(resolved_url) == product_id

    settings = Settings(data_dir=tmp_path)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))
    response = client.post("/api/product-id/extract", json={"share_text": resolved_url})
    assert response.status_code == 200
    assert response.json() == {"product_id": product_id}

    page_html = client.get("/").text
    assert "从分享链接提取商品 ID" in page_html
    assert 'onpaste="scheduleShareProductIdExtraction()"' in page_html
    assert "正在识别商品 ID" in page_html


def test_web_health_and_fixture_collection(tmp_path: Path) -> None:
    settings = Settings(data_dir=tmp_path, min_action_interval=0.1, max_scrolls=20)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))
    assert client.get("/api/health").status_code == 200
    response = client.post(
        "/api/collect",
        json={"keyword": "鸭鸭童装旗舰店", "max_products": 1},
    )
    assert response.status_code == 200
    # FastAPI BackgroundTasks run before the TestClient response is returned.
    job_id = response.json()["job_id"]
    jobs_payload = client.get("/api/jobs").json()
    assert jobs_payload["page"] == 1
    assert jobs_payload["page_size"] == 5
    jobs = jobs_payload["items"]
    assert jobs
    assert any(job["keyword"] == "鸭鸭童装旗舰店" for job in jobs)
    assert any(job.get("product_count") == 1 for job in jobs)
    result = client.get(f"/api/jobs/{job_id}")
    assert result.status_code == 200
    assert result.json()["status"] == "succeeded"
    assert client.post(f"/api/jobs/{job_id}/stop").status_code == 409
    assert client.get(f"/api/jobs/{job_id}/csv").status_code == 200
    assert client.get(f"/api/jobs/{job_id}/xlsx").status_code == 200
    batch = client.get(f"/api/exports/xlsx?job_ids={job_id}")
    assert batch.status_code == 200
    assert batch.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    online = client.get(f"/api/exports/view?job_ids={job_id}")
    assert online.status_code == 200
    assert "鸭鸭童装旗舰店" in online.text
    assert "查询条件：未命名查询条件" in online.text
    assert "店铺详细信息" in online.text
    assert "商品详细信息" in online.text
    assert "列表格式" in online.text
    assert "详细格式" in online.text
    assert "下载 Excel" in online.text
    assert f'href="/api/exports/xlsx?job_ids={job_id}"' in online.text
    assert '<main id="listView">' in online.text
    assert '<main id="detailView" hidden>' in online.text
    assert "<th>图片</th><th>商品名称</th><th>价格</th><th>销量</th>" in online.text
    assert "精准命中状态" not in online.text
    stored = DataStore(settings.database_path).get_result(job_id)
    assert stored and stored.evidence_dir
    image_path = Path(stored.evidence_dir) / "product-test.jpg"
    Image.new("RGB", (20, 20), "red").save(image_path)
    stored.products[0].image_path = image_path.name
    DataStore(settings.database_path).save_result(stored)
    assert client.get(f"/api/jobs/{job_id}/image/{image_path.name}").status_code == 200
    assert client.get(f"/api/jobs/{job_id}/image/../wen.sqlite3").status_code == 404

    second = client.post(
        "/api/collect",
        json={
            "keyword": "鸭鸭童装旗舰店",
            "max_products": 1,
            "query_group_name": "销量前 1 条",
        },
    )
    assert second.status_code == 200
    second_id = second.json()["job_id"]
    combined = client.get(f"/api/exports/xlsx?job_ids={job_id},{second_id}")
    workbook = load_workbook(BytesIO(combined.content))
    assert workbook.sheetnames == ["查询结果", "汇总"]
    result_values = [value for row in workbook["查询结果"].iter_rows(values_only=True) for value in row]
    assert "查询条件：销量前 1 条" in result_values
    assert "商品名称" in result_values
    assert "位置" not in result_values
    assert workbook["汇总"].max_row == 3

    obsolete_precise = client.post(
        "/api/collect",
        json={
            "store_name": "鸭鸭童装旗舰店",
            "max_products": 500,
            "selection_mode": "precise",
            "product_titles": ["已经改名的保存商品"],
            "query_group_name": "精准重点",
        },
    )
    assert obsolete_precise.status_code == 422
    assert "商品 ID" in obsolete_precise.json()["detail"]


def test_query_groups_are_persistent_and_store_name_is_single_input(tmp_path: Path) -> None:
    settings = Settings(data_dir=tmp_path, min_action_interval=0.1, max_scrolls=20)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))
    response = client.post(
        "/api/query-groups",
        json={
            "name": "春季主推",
            "store_name": "鸭鸭童装官方旗舰店",
            "selection_mode": "range",
        },
    )
    assert response.status_code == 200
    group = response.json()
    assert group["store_name"] == "鸭鸭童装官方旗舰店"
    assert client.get("/api/query-groups").json()[0]["name"] == "春季主推"
    page_html = client.get("/").text
    assert "春季主推" in page_html
    assert "stopQueryButton" in page_html
    assert "/api/jobs/" in page_html and "/stop" in page_html
    assert page_html.index('id="queryProgress"') < page_html.index('id="stopQueryButton"')
    assert page_html.index('id="catalogChoose"') < page_html.index('id="catalogStop"')
    assert "void stopCatalogRead()" in page_html
    assert "任务仍然存在，请再次停止" in page_html
    assert "cleanup_timeout" not in page_html
    # The initial response is synchronous, while mutations refresh the saved
    # list from the API so a newly-created condition appears immediately.
    assert "async function refreshGroups()" in page_html
    assert "await refreshGroups()" in page_html
    plan = client.post(
        "/api/query-plans",
        json={"name": "每日童装监测", "group_ids": [group["id"]]},
    )
    assert plan.status_code == 200
    assert client.get(f"/api/query-plans/{plan.json()['id']}").json()["group_ids"] == [group["id"]]

    unnamed = client.post(
        "/api/query-groups",
        json={"store_name": "另一个店铺", "selection_mode": "range"},
    )
    assert unnamed.status_code == 200
    assert unnamed.json()["name"] == "查询店铺内的商品 · 另一个店铺"
    listed = client.get("/api/query-groups").json()
    assert [item["store_name"] for item in listed] == [
        "鸭鸭童装官方旗舰店",
        "另一个店铺",
    ]
    assert listed[-1]["name"] == "查询店铺内的商品 · 另一个店铺"
    renamed = client.patch(
        f"/api/query-groups/{unnamed.json()['id']}/name",
        json={"name": "另一个店铺重点款"},
    )
    assert renamed.status_code == 200
    assert renamed.json()["name"] == "另一个店铺重点款"

    collected = client.post(
        "/api/collect",
        json={
            "store_name": "鸭鸭童装旗舰店",
            "max_products": 1,
            "product_titles": [f"商品{i}" for i in range(25)],
        },
    )
    assert collected.status_code == 200
    assert client.get(f"/api/jobs/{collected.json()['job_id']}").status_code == 200


def test_query_groups_support_product_ids_and_sec_shop_id(tmp_path: Path) -> None:
    settings = Settings(data_dir=tmp_path, min_action_interval=0.1)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))

    product_ids = client.post(
        "/api/query-groups",
        json={
            "selection_mode": "precise",
            # precise_query_mode intentionally omitted: precise is always ID-based.
            "store_name": "旧页面残留的店铺名",
            "product_ids": ["3817041292878283160", "3832268942550892829"],
        },
    )
    assert product_ids.status_code == 200
    assert product_ids.json()["store_name"] is None
    assert product_ids.json()["precise_query_mode"] == "product_ids"
    assert product_ids.json()["name"] == "精准查询·商品 ID"

    direct_store = client.post(
        "/api/query-groups",
        json={
            "selection_mode": "range",
            "store_locator_mode": "sec_shop_id",
            "sec_shop_id": "njBWlyXQ",
            "limit_count": 10,
        },
    )
    assert direct_store.status_code == 200
    assert direct_store.json()["store_name"] is None
    assert direct_store.json()["sec_shop_id"] == "njBWlyXQ"
    assert direct_store.json()["name"] == "查询店铺内的商品 · njBWlyXQ"

    invalid = client.post(
        "/api/query-groups",
        json={
            "selection_mode": "precise",
            "precise_query_mode": "product_ids",
            "product_ids": ["not-an-id"],
        },
    )
    assert invalid.status_code == 422
    page = client.get("/").text
    assert "查询我收藏的商品" in page
    assert "查询店铺内的商品" in page
    assert page.index('<option value="precise" selected>') < page.index(
        '<option value="favorites">'
    )
    assert "还没有保存的查询条件" not in page
    assert '<section class="card" id="historyCard" hidden>' in page
    assert "还没有查询记录" not in page
    assert "根据店铺查询" not in page
    assert "可用逗号或换行分隔" in page
    assert "sec_shop_id（直达）" in page


def test_query_runs_group_conditions_and_use_one_daily_xlsx_name(tmp_path: Path) -> None:
    settings = Settings(data_dir=tmp_path, min_action_interval=0.1, max_scrolls=20)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))
    run_id = "query-run-test"
    job_ids = []
    for name in ("综合前 1 条", "销量前 1 条"):
        response = client.post(
            "/api/collect",
            json={
                "store_name": "鸭鸭童装旗舰店",
                "max_products": 1,
                "query_run_id": run_id,
                "query_group_name": name,
            },
        )
        assert response.status_code == 200
        job_ids.append(response.json()["job_id"])

    runs = client.get("/api/query-runs").json()
    assert runs["total"] == 1
    assert runs["items"][0]["id"] == run_id
    assert runs["items"][0]["job_ids"] == job_ids
    assert [group["query_group_name"] for group in runs["items"][0]["groups"]] == [
        "综合前 1 条", "销量前 1 条"
    ]
    exported = client.get(f"/api/exports/xlsx?job_ids={','.join(job_ids)}")
    assert exported.status_code == 200
    assert "wen-dy-" in exported.headers["content-disposition"]

    evidence_dirs = [
        Path(DataStore(settings.database_path).get_result(job_id).evidence_dir)
        for job_id in job_ids
    ]
    assert all(path.exists() for path in evidence_dirs)
    deleted = client.delete(f"/api/query-runs/{run_id}")
    assert deleted.status_code == 200
    assert deleted.json()["deleted_jobs"] == 2
    assert client.get("/api/query-runs").json()["total"] == 0
    assert all(not path.exists() for path in evidence_dirs)
    assert client.delete(f"/api/query-runs/{run_id}").status_code == 404


def test_precise_catalog_jobs_are_not_query_history_or_snapshots(tmp_path: Path) -> None:
    settings = Settings(data_dir=tmp_path, min_action_interval=0.1, max_scrolls=20)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))
    response = client.post(
        "/api/collect",
        json={
            "store_name": "鸭鸭童装旗舰店",
            "max_products": 500,
            "selection_mode": "precise_catalog",
        },
    )
    assert response.status_code == 200
    job_id = response.json()["job_id"]
    result = client.get(f"/api/jobs/{job_id}").json()
    assert result["status"] == "succeeded"
    assert result["selection_mode"] == "precise_catalog"
    assert client.get("/api/query-runs").json()["total"] == 0
    assert client.get("/api/jobs").json()["total"] == 0
    # The result is retained for the picker/polling endpoint, but catalog-only
    # work must not pollute historical snapshots.
    assert DataStore(settings.database_path).get_result(job_id) is not None
    with sqlite3.connect(settings.database_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM store_snapshots").fetchone()[0] == 0
        assert connection.execute("SELECT COUNT(*) FROM product_snapshots").fetchone()[0] == 0


def test_favorites_query_group_and_collection_do_not_require_store(tmp_path: Path) -> None:
    settings = Settings(data_dir=tmp_path, min_action_interval=0.1, max_scrolls=20)
    client = TestClient(create_app(settings, device_factory=lambda _: FixtureDevice()))
    group_response = client.post(
        "/api/query-groups",
        json={"selection_mode": "favorites", "name": "我的收藏全部商品"},
    )
    assert group_response.status_code == 200
    group = group_response.json()
    assert group["store_name"] is None
    assert group["selection_mode"] == "favorites"

    collect_response = client.post(
        "/api/collect",
        json={
            "selection_mode": "favorites",
            "max_products": 500,
            "query_group_id": group["id"],
            "query_group_name": group["name"],
            "query_run_id": "favorites-run",
        },
    )
    assert collect_response.status_code == 200
    job_id = collect_response.json()["job_id"]
    result = client.get(f"/api/jobs/{job_id}").json()
    assert result["selection_mode"] == "favorites"
    assert result["requested_store_name"] is None
    assert result["store"] is None
    runs = client.get("/api/query-runs").json()["items"]
    assert runs[0]["groups"][0]["store_name"] == "我的收藏"
    xlsx = client.get(f"/api/exports/xlsx?job_ids={job_id}")
    workbook = load_workbook(BytesIO(xlsx.content))
    values = [value for row in workbook["查询结果"].iter_rows(values_only=True) for value in row]
    assert "我的收藏" in values
